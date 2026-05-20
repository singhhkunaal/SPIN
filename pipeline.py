"""
Treya Spend Diagnostic — analysis pipeline.

Refactor of the original script (script.py) for web deployment:
  * Reads the uploaded workbook from a file-like object instead of a Windows path.
  * Persistent caches live in Supabase, not local JSON files.
  * All progress messages go through a `status` callback (Streamlit-friendly).
  * Returns the populated workbook as in-memory bytes for download.

Bug fix from original (lines 981–986): the dedupe-enrichment loop wrote results
to `vendor_map_memory` with `area_mapping`/`category_mapping` keys, which are
fields that loop never produces. It should have written to `vendor_memory` with
`standard_name`/`parent`/`notes`/`confidence_score`. Fixed here.
"""

from __future__ import annotations
import io
import json
import re
import time
from collections import defaultdict, Counter
from datetime import datetime
from typing import Callable, Optional

import anthropic
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

import cleaning
import database as db

# ---------------------------------------------------------------------------
# Pipeline constants (lifted from the original script)
# ---------------------------------------------------------------------------
BATCH_SIZE = 25
AI_TAXONOMY_CEILING = 15
MAX_RETRIES = 3
RETRY_WAIT = 60

SAVINGS_THRESHOLD = 0.80
SAVINGS_MIN_CATEGORIES = 5

DARK_BLUE = "002060"
WHITE = "FFFFFF"
MID_GREY = "404040"


# ---------------------------------------------------------------------------
# Status reporting — defaults to print if no callback supplied.
# ---------------------------------------------------------------------------
StatusFn = Callable[[str], None]


def _noop_status(msg: str) -> None:
    print(msg)


# ---------------------------------------------------------------------------
# Formatting helpers (verbatim from original script)
# ---------------------------------------------------------------------------
def set_sheet_view(ws, zoom=80, gridlines=False):
    ws.sheet_view.showGridLines = gridlines
    ws.sheet_view.zoomScale = zoom


def style_heading(cell, size=11, bold=True, color="000000"):
    cell.font = Font(name="Calibri", bold=bold, size=size, color=color)


def style_subheading(cell):
    cell.font = Font(name="Calibri", bold=True, size=10, color=MID_GREY)


def style_body(cell, wrap=False, size=10):
    cell.font = Font(name="Calibri", bold=False, size=size, color="000000")
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def style_table_header(cell):
    cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    cell.fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)


def style_data_row(cell):
    cell.font = Font(name="Calibri", bold=False, size=10, color="000000")


def style_total_row(cell):
    cell.font = Font(name="Calibri", bold=True, size=10, color="000000")


def clean_research_text(text: str) -> str:
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'#{1,6}\s*', '', text)
    text = re.sub(r'---+', '', text)
    text = re.sub(r'^\s*[-•]\s*', '', text, flags=re.MULTILINE)
    return text.strip()


# ---------------------------------------------------------------------------
# Claude wrapper
# ---------------------------------------------------------------------------
def _call_claude_with_retry(client: anthropic.Anthropic, kwargs: dict,
                            label: str, status: StatusFn):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return client.messages.create(**kwargs)
        except Exception as e:
            if "429" in str(e):
                wait = RETRY_WAIT * attempt
                status(f"  ⏳ Rate limit hit ({label}) — waiting {wait}s before retry {attempt}/{MAX_RETRIES}…")
                time.sleep(wait)
            else:
                status(f"  ⚠️ API error ({label}): {e}")
                return None
    status(f"  ⚠️ All retries exhausted for {label}")
    return None


def _extract_json_array(text: str):
    json_start = text.find("[")
    json_end = text.rfind("]") + 1
    if json_start == -1 or json_end == 0:
        return None
    try:
        return json.loads(text[json_start:json_end])
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Claude steps
# ---------------------------------------------------------------------------
def research_company(client, supabase, company, pe, ind, status: StatusFn) -> str:
    cache_key = f"{company}|{pe}"
    cached = db.load_research(supabase, cache_key)
    if cached:
        status("  ↩️  Using cached research.")
        return cached

    prompt = f"""
You are a procurement and business research analyst.

Research the following company and provide a structured summary:

Company Name: {company}
PE Firm (Owner): {pe if pe else "Not specified"}
Known Industry: {ind if ind != "Unknown" else "Unknown - please identify"}

Use web search to find accurate, current information. Provide your findings in plain text
with NO markdown formatting, NO asterisks, NO hash symbols, NO bullet dashes.
Use numbered sections only.

1. COMPANY OVERVIEW
Brief description of what the company does (2-3 sentences).

2. INDUSTRY AND SECTOR
Primary industry, sector, and relevant sub-sectors.

3. BUSINESS MODEL
How the company makes money. Product-based, service-based, or mixed?

4. KEY PRODUCTS AND SERVICES
Main offerings (max 6 items, plain numbered list).

5. DIRECT COST CATEGORIES (COGS)
Likely direct cost / COGS spend categories for this type of business (max 6 items).

6. LIKELY INDIRECT SPEND AREAS
Key indirect spend categories relevant to this company (max 8 items).

7. PE FIRM CONTEXT
Relevant context about the PE firm portfolio focus or strategy if known.
Include any known portfolio companies under this PE firm as they may appear as vendors.

8. TAXONOMY RECOMMENDATIONS
Suggest 5-10 custom spend categories specifically relevant to this company.
Format each as: Area | Category

Keep each section factual and concise. No markdown formatting whatsoever.
"""
    response = _call_claude_with_retry(client, {
        "model": "claude-sonnet-4-6",
        "max_tokens": 3000,
        "tools": [{"type": "web_search_20250305", "name": "web_search"}],
        "messages": [{"role": "user", "content": prompt}],
    }, label="research", status=status)

    if response is None:
        fallback = f"Research could not be completed for {company}."
        db.save_research(supabase, cache_key, fallback)
        return fallback

    research_text = "".join(b.text for b in response.content if getattr(b, "type", None) == "text")
    if not research_text.strip():
        research_text = f"Research could not be completed for {company}. Please populate manually."
    research_text = clean_research_text(research_text)
    db.save_research(supabase, cache_key, research_text)
    return research_text


def enrich_dedupe_batch(client, vendor_batch, canonical_spend,
                        canonical_descriptions, status: StatusFn) -> list:
    vendor_details = []
    for v in vendor_batch:
        entry = {"vendor": v}
        spend = canonical_spend.get(v, 0)
        if spend:
            entry["total_spend"] = round(spend, 2)
        descs = canonical_descriptions.get(v, [])
        if descs:
            entry["invoice_descriptions"] = descs
        vendor_details.append(entry)

    prompt = f"""
You are a vendor data standardization expert.

For each vendor below, return a clean standardized name, parent company, QC note and confidence score.

CRITICAL RULES:
1. standard_name: Clean, properly formatted version of the vendor name
2. parent: ONLY populate if the vendor is a KNOWN subsidiary of a larger company
   (e.g. Whole Foods → Amazon, Instagram → Meta, YouTube → Alphabet)
   If you are not certain the vendor is a subsidiary, leave parent as an EMPTY STRING ""
   Do NOT repeat the standard_name as the parent
   Do NOT guess or infer a parent — only use it when you are certain
3. notes: 1-2 sentences explaining your standardization decision for QC
4. confidence_score: "High", "Medium", "Low", or "Requires Review"

Return ONLY a valid JSON array in the EXACT same order as the input vendors list.
The array must have exactly {len(vendor_batch)} items — one per vendor.

[
  {{
    "standard_name": "<cleaned name>",
    "parent": "<parent company or empty string>",
    "notes": "<brief QC note>",
    "confidence_score": "<High|Medium|Low|Requires Review>"
  }}
]

Vendors:
{json.dumps(vendor_details, indent=2)}
"""
    response = _call_claude_with_retry(client, {
        "model": "claude-sonnet-4-6",
        "max_tokens": 8192,
        "messages": [{"role": "user", "content": prompt}],
    }, label="dedupe", status=status)

    if response is None:
        return []
    text = "".join(b.text for b in response.content if getattr(b, "type", None) == "text")
    return _extract_json_array(text) or []


def generate_ai_taxonomy(client, research_summary, deduped_names, canonical_spend,
                         taxonomy_text, company_name, industry, pe_firm,
                         status: StatusFn) -> list:
    vendor_context = []
    for v in deduped_names[:100]:
        entry = {"vendor": v}
        spend = canonical_spend.get(v, 0)
        if spend:
            entry["total_spend"] = round(spend, 2)
        vendor_context.append(entry)

    prompt = f"""
You are a procurement taxonomy specialist.

Company: {company_name} | Industry: {industry} | PE Firm: {pe_firm or "Not specified"}
Total unique vendors: {len(deduped_names)}

Research Summary:
{research_summary[:1500]}

Existing Standard Taxonomy (do NOT duplicate):
{taxonomy_text}

Vendor list with spend where available:
{json.dumps(vendor_context, indent=2)}

Generate ADDITIONAL spend categories that are:
1. Specifically relevant to this company's industry, business model, and vendor base
2. NOT already in the standard taxonomy
3. Only if 2+ vendors would logically fall under it
4. Maximum {AI_TAXONOMY_CEILING} categories, minimum 3
5. Use existing Areas where possible
6. Do not include Intra-company Transfer
7. For any direct cost or COGS category, the area MUST be exactly "COGS"

Return ONLY a valid JSON array, no extra text.
[
  {{"area": "<Area>", "category": "<Category>"}}
]
"""
    response = _call_claude_with_retry(client, {
        "model": "claude-sonnet-4-6",
        "max_tokens": 1500,
        "messages": [{"role": "user", "content": prompt}],
    }, label="taxonomy", status=status)

    if response is None:
        return []
    text = "".join(b.text for b in response.content if getattr(b, "type", None) == "text")
    parsed = _extract_json_array(text) or []
    return parsed[:AI_TAXONOMY_CEILING]


def enrich_map_batch(client, vendor_batch, research_summary, enriched_taxonomy_text,
                     canonical_spend, company_name, industry, pe_firm,
                     status: StatusFn) -> list:
    vendor_details = []
    for v in vendor_batch:
        entry = {"vendor": v}
        spend = canonical_spend.get(v, 0)
        if spend:
            entry["total_spend"] = round(spend, 2)
        vendor_details.append(entry)

    prompt = f"""
You are an expert procurement analyst specializing in spend categorization.

Company: {company_name}
PE Firm: {pe_firm or "Not specified"}
Industry: {industry}

Company Research Summary:
{research_summary[:2000]}

Full Taxonomy (Standard + AI Generated - use ALL of these when mapping):
{enriched_taxonomy_text}

INTRA-COMPANY DETECTION RULES:
For each vendor, assess whether it could be a related party or intra-company transaction
by cross-referencing against the company name, PE firm, known subsidiaries, and portfolio
companies mentioned in the research summary.

THREE-TIER APPROACH:
TIER 1 - CLEARLY INTRA-COMPANY:
Assign area_mapping="Unaddressable", category_mapping="Intra-company Transfer",
confidence_score="High", and explain in notes.

TIER 2 - POSSIBLY INTRA-COMPANY:
Use normal mapping BUT confidence_score="Requires Review" and flag suspicion in notes.

TIER 3 - CLEARLY NOT INTRA-COMPANY:
Apply normal mapping rules below.

NORMAL MAPPING RULES:
1. Assign EXACTLY ONE area and ONE category per vendor
2. For ANY vendor supplying direct materials, raw materials, ingredients, packaging materials,
   or core product costs, you MUST set area_mapping to exactly "COGS" — no other value is acceptable.
3. Use the taxonomy above for all indirect vendors
4. Only create a brand new category if absolutely nothing in the full taxonomy fits
5. Never leave area or category blank
6. Use your world knowledge of what each vendor does — do not guess based on vendor name alone

Return ONLY a valid JSON array in the EXACT same order as the input vendors list.
The array must have exactly {len(vendor_batch)} items — one per vendor.

[
  {{
    "area_mapping": "<exactly one Area>",
    "category_mapping": "<exactly one Category>",
    "notes": "<1-2 sentence QC note>",
    "confidence_score": "<High|Medium|Low|Requires Review>"
  }}
]

Vendors:
{json.dumps(vendor_details, indent=2)}
"""
    response = _call_claude_with_retry(client, {
        "model": "claude-sonnet-4-6",
        "max_tokens": 2000,
        "messages": [{"role": "user", "content": prompt}],
    }, label="map", status=status)

    if response is None:
        return []
    text = "".join(b.text for b in response.content if getattr(b, "type", None) == "text")
    return _extract_json_array(text) or []


def get_savings_benchmarks(client, categories, research_summary, savings_cache,
                           company_name, industry, pe_firm,
                           status: StatusFn) -> tuple[dict, dict]:
    """Returns (all_benchmarks, newly_fetched_only). The caller persists the
    new ones to Supabase; the merged dict is what the workbook reads from."""
    results: dict = {}
    new_entries: dict = {}
    to_lookup = [cat for cat in categories if cat not in savings_cache]
    for cat in categories:
        if cat in savings_cache:
            results[cat] = savings_cache[cat]
    if not to_lookup:
        return results, new_entries

    prompt = f"""
You are a procurement savings benchmarking expert.

Company: {company_name} | Industry: {industry} | PE Firm: {pe_firm or "Not specified"}

Research Context:
{research_summary[:1000]}

For each spend category below, provide industry-standard savings benchmarks AND
the sourcing strategy needed to realise them.

1. addressability: % of spend typically addressable/sourceable (0.0-1.0)
   - COGS categories: typically 0.3-0.6
   - Indirect categories: typically 0.6-0.9
2. savings_low_pct:  Conservative savings % achievable (0.0-1.0)
3. savings_high_pct: Aggressive savings % achievable (0.0-1.0)
4. notes: 1-2 sentences explaining benchmark rationale and industry context.
5. strategy: An object with:
   - headline: short, punchy 5-10 word strategy summary (e.g. "Rebroke, redesign, renegotiate carrier terms")
   - levers: array of 3-5 concrete sourcing levers (each a single sentence, action-oriented)
   - timeline: short string describing realistic timing (e.g. "90-120 days through competitive selection")

Return ONLY a valid JSON array, no extra text.
[
  {{
    "category": "<category name>",
    "addressability": 0.0,
    "savings_low_pct": 0.0,
    "savings_high_pct": 0.0,
    "notes": "<rationale>",
    "strategy": {{
      "headline": "<5-10 word summary>",
      "levers":   ["<lever 1>", "<lever 2>", "<lever 3>"],
      "timeline": "<timing string>"
    }}
  }}
]

Categories:
{json.dumps(to_lookup)}
"""
    response = _call_claude_with_retry(client, {
        "model": "claude-sonnet-4-6",
        "max_tokens": 2000,
        "messages": [{"role": "user", "content": prompt}],
    }, label="benchmarks", status=status)

    if response is None:
        return results, new_entries

    text = "".join(b.text for b in response.content if getattr(b, "type", None) == "text")
    batch_results = _extract_json_array(text) or []
    for item in batch_results:
        cat = item.get("category")
        if not cat:
            continue
        bench = {
            "addressability": item.get("addressability", 0.8),
            "savings_low_pct": item.get("savings_low_pct", 0.01),
            "savings_high_pct": item.get("savings_high_pct", 0.02),
            "notes": item.get("notes", ""),
            "strategy": item.get("strategy") or {},
        }
        results[cat] = bench
        new_entries[cat] = bench
    return results, new_entries


# ---------------------------------------------------------------------------
# Top categories selection
# ---------------------------------------------------------------------------
def _select_top_categories(sorted_cats, status: StatusFn):
    excluded = {
        "Unaddressable", "Unmapped", "Intra-company Transfer",
        "(blank)", "-", "", "Grand Total", "Requires Review",
    }
    filtered = [
        (cat, spend) for cat, spend in sorted_cats
        if cat not in excluded and not cat.lower().startswith("unaddressable")
    ]
    total_filtered = sum(s for _, s in filtered)
    top, cumulative = [], 0
    for cat, spend in filtered:
        cumulative += spend
        top.append((cat, spend))
        if total_filtered > 0 and cumulative / total_filtered >= SAVINGS_THRESHOLD:
            break
    if len(top) < SAVINGS_MIN_CATEGORIES:
        already = set(c for c, _ in top)
        remaining = [(c, s) for c, s in filtered if c not in already]
        needed = SAVINGS_MIN_CATEGORIES - len(top)
        top.extend(remaining[:needed])
    if len(top) < SAVINGS_MIN_CATEGORIES:
        status(f"  ⚠️ Only {len(top)} addressable categories available "
               f"(minimum is {SAVINGS_MIN_CATEGORIES}). Using all available.")
    pct = (sum(s for _, s in top) / total_filtered * 100) if total_filtered > 0 else 0
    status(f"  ✅ {len(top)} categories selected — covering {pct:.1f}% of addressable spend")
    return top


# ---------------------------------------------------------------------------
# THE PIPELINE
# ---------------------------------------------------------------------------
def run_pipeline(
    uploaded_workbook_bytes: bytes,
    anthropic_api_key: str,
    supabase_url: Optional[str] = None,
    supabase_key: Optional[str] = None,
    status: StatusFn = _noop_status,
) -> tuple[bytes, str, dict]:
    """Run the full diagnostic pipeline.

    Returns (output_xlsx_bytes, suggested_filename, dashboard_data).
    dashboard_data is the IRONCLAD-shaped dict (see dashboard_export.py).
    """

    # ----- Clients -----
    client = anthropic.Anthropic(api_key=anthropic_api_key)
    supabase = None
    if supabase_url and supabase_key:
        try:
            supabase = db.get_client(supabase_url, supabase_key)
        except Exception as e:
            status(f"  ⚠️ Could not connect to Supabase ({e}). Continuing without cache.")
            supabase = None

    # ----- Load caches -----
    status("📦 Loading caches from Supabase…")
    vendor_memory = db.load_vendor_dedupe_memory(supabase)
    vendor_map_memory = db.load_vendor_map_memory(supabase)
    savings_cache = db.load_savings_benchmarks(supabase)
    status(f"  ✅ Cache: {len(vendor_memory)} dedupe entries · {len(vendor_map_memory)} map entries · {len(savings_cache)} benchmarks")

    # ----- Load workbook from uploaded bytes -----
    wb = load_workbook(io.BytesIO(uploaded_workbook_bytes))
    ws_data_tab = wb["Data"]
    ws_vendor = wb["Vendor Dedupe"]
    ws_map = wb["Vendor Map"]
    ws_tax = wb["Taxonomy"]
    ws_notes = wb["Notes"]
    ws_cd = wb["Consolidated Data"]
    ws_sbc = wb["Spend by Category"]
    ws_sav = wb["Savings Opportunity"]

    company_name = ws_notes["D3"].value or ws_cd["B2"].value
    pe_firm = ws_notes["D4"].value or ""
    industry = ws_notes["D5"].value or "Unknown"

    status(f"🏢 Company : {company_name}")
    status(f"🏦 PE Firm : {pe_firm or 'Not specified'}")
    status(f"🏭 Industry: {industry}")

    # ----- Extract taxonomy -----
    taxonomy = []
    for row in ws_tax.iter_rows(min_row=6, min_col=2, max_col=3, values_only=True):
        area, category = row
        if area and category and area not in ("Area", "Category", "Added Categories", "Custom Category"):
            taxonomy.append({"area": area, "category": category})

    def build_taxonomy_text(taxonomy_list):
        return "\n".join([f"  - Area: {t['area']} | Category: {t['category']}" for t in taxonomy_list])

    taxonomy_text = build_taxonomy_text(taxonomy)

    # ----- Extract vendors / spend / descriptions -----
    vendor_spend = defaultdict(float)
    vendor_descriptions = defaultdict(list)
    vendor_rows = []
    for row in ws_cd.iter_rows(min_row=6, min_col=2, max_col=9, values_only=True):
        _, _, vendor, invoice_desc, _, _, _, spend = row
        if not vendor:
            continue
        vendor = str(vendor).strip()
        vendor_rows.append(vendor)
        if spend and isinstance(spend, (int, float)):
            vendor_spend[vendor] += spend
        if invoice_desc and str(invoice_desc).strip():
            vendor_descriptions[vendor].append(str(invoice_desc).strip())

    unique_vendors = list(set(vendor_rows))
    vendor_top_descriptions = {v: [d for d, _ in Counter(ds).most_common(3)]
                               for v, ds in vendor_descriptions.items()}
    total_raw_spend = sum(vendor_spend.values())

    # ----- V9 cleaning + fuzzy grouping -----
    status(f"\n🧹 Cleaning {len(unique_vendors)} unique vendors…")
    cleaned_vendors_v9 = cleaning.clean_batch_v9(unique_vendors)
    v9_clean_map = dict(zip(unique_vendors, cleaned_vendors_v9))
    cleaned_vendor_list = [v9_clean_map.get(v, cleaning.clean_single(v)) for v in unique_vendors]
    vendor_mapping = cleaning.group_vendors(cleaned_vendor_list)
    status("  ✅ Cleaning + fuzzy grouping complete.")

    def get_canonical(raw_vendor):
        v9_cleaned = v9_clean_map.get(raw_vendor, cleaning.clean_single(raw_vendor))
        return vendor_mapping.get(v9_cleaned, v9_cleaned)

    canonical_spend = defaultdict(float)
    canonical_descriptions = defaultdict(list)
    for raw_vendor in unique_vendors:
        canonical = get_canonical(raw_vendor)
        canonical_spend[canonical] += vendor_spend.get(raw_vendor, 0)
        canonical_descriptions[canonical].extend(vendor_top_descriptions.get(raw_vendor, []))
    for canonical in canonical_descriptions:
        canonical_descriptions[canonical] = list(dict.fromkeys(canonical_descriptions[canonical]))[:3]

    # ----- STEP 1: Research -----
    status(f"\n🔍 Researching {company_name}…")
    research_summary = research_company(client, supabase, company_name, pe_firm, industry, status)
    status("✅ Research complete.")

    # ----- STEP 2: Dedupe enrichment (BUG FIX vs original) -----
    unique_canonicals = list(set(vendor_mapping.values()))
    to_enrich = sorted([v for v in unique_canonicals if v not in vendor_memory])
    status(f"\n📦 Vendors to dedupe-enrich: {len(to_enrich)} "
           f"(skipping {len(unique_canonicals) - len(to_enrich)} cached)")

    for i in range(0, len(to_enrich), BATCH_SIZE):
        batch = to_enrich[i:i + BATCH_SIZE]
        status(f"  🔄 Dedupe batch {i // BATCH_SIZE + 1}/{-(-len(to_enrich) // BATCH_SIZE)} ({len(batch)} vendors)…")
        results = enrich_dedupe_batch(client, batch, canonical_spend, canonical_descriptions, status)

        new_batch = {}
        for idx, original_vendor in enumerate(batch):
            if idx >= len(results):
                continue
            item = results[idx]
            entry = {
                "standard_name": item.get("standard_name", original_vendor),
                "parent": (item.get("parent") or "").strip(),
                "notes": item.get("notes", ""),
                "confidence_score": item.get("confidence_score", "Requires Review"),
            }
            vendor_memory[original_vendor] = entry
            new_batch[original_vendor] = entry

        db.save_vendor_dedupe_batch(supabase, new_batch)

    # Write Vendor Dedupe tab
    start_row = 7
    for i, raw_vendor in enumerate(unique_vendors):
        canonical = get_canonical(raw_vendor)
        enriched = vendor_memory.get(canonical, {
            "standard_name": canonical, "parent": "",
            "notes": "", "confidence_score": "Requires Review",
        })
        r = start_row + i
        std_name = enriched.get("standard_name", canonical)
        parent = enriched.get("parent", "").strip()
        if parent.upper() == std_name.upper():
            parent = ""
        ws_vendor[f"B{r}"] = raw_vendor.strip()
        ws_vendor[f"C{r}"] = f"=SUMIFS(Data!G:G,Data!C:C,'Vendor Dedupe'!B{r})"
        ws_vendor[f"D{r}"] = std_name
        ws_vendor[f"E{r}"] = parent
        ws_vendor[f"F{r}"] = enriched.get("notes", "")
        ws_vendor[f"G{r}"] = enriched.get("confidence_score", "Requires Review")
    set_sheet_view(ws_vendor)
    status("✅ Vendor Dedupe tab populated.")

    # Build deduped names list
    deduped_names, seen = [], set()
    for raw_vendor in unique_vendors:
        canonical = get_canonical(raw_vendor)
        enriched = vendor_memory.get(canonical, {"standard_name": canonical})
        std_name = enriched.get("standard_name", canonical)
        if std_name not in seen:
            seen.add(std_name)
            deduped_names.append(std_name)

    # ----- STEP 3: AI Taxonomy -----
    status("\n🧠 Generating AI taxonomy…")
    ai_taxonomy = generate_ai_taxonomy(client, research_summary, deduped_names, canonical_spend,
                                       taxonomy_text, company_name, industry, pe_firm, status)
    status(f"  ✅ {len(ai_taxonomy)} custom categories generated.")

    tax_last_row = ws_tax.max_row
    ai_heading_row = tax_last_row + 3
    ws_tax[f"B{ai_heading_row}"].value = "AI Generated Taxonomy"
    style_heading(ws_tax[f"B{ai_heading_row}"], size=11, bold=True)
    sub_row = ai_heading_row + 1
    ws_tax[f"B{sub_row}"].value = "Area"
    ws_tax[f"C{sub_row}"].value = "Category"
    style_subheading(ws_tax[f"B{sub_row}"])
    style_subheading(ws_tax[f"C{sub_row}"])
    for j, item in enumerate(ai_taxonomy):
        r = sub_row + 1 + j
        ws_tax[f"B{r}"].value = item.get("area", "")
        ws_tax[f"C{r}"].value = item.get("category", "")
        style_body(ws_tax[f"B{r}"])
        style_body(ws_tax[f"C{r}"])
    set_sheet_view(ws_tax)

    enriched_taxonomy = taxonomy + ai_taxonomy
    enriched_taxonomy_text = build_taxonomy_text(enriched_taxonomy)

    # ----- STEP 4: Vendor Map enrichment -----
    to_map = sorted([v for v in deduped_names if v not in vendor_map_memory])
    status(f"\n🗺️  Vendors to map: {len(to_map)} (skipping {len(deduped_names) - len(to_map)} cached)")

    for i in range(0, len(to_map), BATCH_SIZE):
        batch = to_map[i:i + BATCH_SIZE]
        status(f"  🔄 Map batch {i // BATCH_SIZE + 1}/{-(-len(to_map) // BATCH_SIZE)} ({len(batch)} vendors)…")
        results = enrich_map_batch(client, batch, research_summary, enriched_taxonomy_text,
                                    canonical_spend, company_name, industry, pe_firm, status)
        new_batch = {}
        for idx, original_vendor in enumerate(batch):
            if idx >= len(results):
                continue
            item = results[idx]
            entry = {
                "area_mapping": item.get("area_mapping", "Requires Review"),
                "category_mapping": item.get("category_mapping", "Requires Review"),
                "notes": item.get("notes", ""),
                "confidence_score": item.get("confidence_score", "Requires Review"),
            }
            vendor_map_memory[original_vendor] = entry
            new_batch[original_vendor] = entry
        db.save_vendor_map_batch(supabase, new_batch)

    # Write Vendor Map tab
    map_start_row = 7
    for i, vendor_name in enumerate(deduped_names):
        mapped = vendor_map_memory.get(vendor_name, {
            "area_mapping": "Requires Review", "category_mapping": "Requires Review",
            "notes": "", "confidence_score": "Requires Review",
        })
        r = map_start_row + i
        ws_map[f"B{r}"] = vendor_name
        ws_map[f"C{r}"] = f"=SUMIFS(Data!G:G,Data!D:D,'Vendor Map'!B{r})"
        ws_map[f"D{r}"] = mapped.get("area_mapping", "Requires Review")
        ws_map[f"E{r}"] = mapped.get("category_mapping", "Requires Review")
        ws_map[f"F{r}"] = mapped.get("notes", "")
        ws_map[f"G{r}"] = mapped.get("confidence_score", "Requires Review")
    set_sheet_view(ws_map)
    status("✅ Vendor Map tab populated.")

    # ----- STEP 5: Company Research tab -----
    status("\n📄 Creating Company Research tab…")
    if "Company Research" in wb.sheetnames:
        del wb["Company Research"]
    ws_research = wb.create_sheet("Company Research", 3)
    set_sheet_view(ws_research)
    ws_research["B2"] = company_name
    style_heading(ws_research["B2"], size=14, bold=True)
    ws_research["B3"] = "Spend and Opportunity Assessment"
    style_heading(ws_research["B3"], size=11, bold=False, color=MID_GREY)
    ws_research["B4"] = "Company Research"
    style_heading(ws_research["B4"], size=11, bold=True)

    meta = [
        ("Company Name", company_name),
        ("PE Firm", pe_firm or "Not specified"),
        ("Industry", industry if industry != "Unknown" else "Not specified"),
        ("Research Date", datetime.today().strftime("%Y-%m-%d")),
    ]
    for col_idx, (hdr, val) in enumerate(meta, start=2):
        hdr_cell = ws_research.cell(row=6, column=col_idx)
        val_cell = ws_research.cell(row=7, column=col_idx)
        hdr_cell.value = hdr
        val_cell.value = val
        style_subheading(hdr_cell)
        style_body(val_cell)

    ws_research["B9"] = "Research Summary"
    style_heading(ws_research["B9"], size=11, bold=True)
    current_row = 10
    for line in research_summary.split("\n"):
        line = line.strip()
        if not line:
            current_row += 1
            continue
        cell = ws_research.cell(row=current_row, column=2)
        cell.value = line
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        is_section_header = bool(re.match(r"^\d+\.", line)) or line.endswith(":")
        if is_section_header:
            cell.font = Font(name="Calibri", bold=True, size=10, color=MID_GREY)
        else:
            style_body(cell, wrap=True)
        current_row += 1

    ws_research.column_dimensions["A"].width = 2.63
    ws_research.column_dimensions["B"].width = 80
    for col in ["C", "D", "E", "F"]:
        ws_research.column_dimensions[col].width = 20

    # ----- STEP 6: Populate Data tab -----
    status("\n📋 Populating Data tab…")
    cd_last_row = 5
    for row in ws_cd.iter_rows(min_row=6, min_col=4, max_col=4):
        if row[0].value is not None:
            cd_last_row = row[0].row
    for row in ws_data_tab.iter_rows(min_row=8, min_col=15, max_col=24):
        for cell in row:
            cell.value = None
    dest_row = 8
    rows_written = 0
    for src_row_idx in range(6, cd_last_row + 1):
        has_data = False
        for col_offset in range(10):
            src_val = ws_cd.cell(row=src_row_idx, column=2 + col_offset).value
            if src_val is not None:
                has_data = True
            ws_data_tab.cell(row=dest_row, column=15 + col_offset).value = src_val
        if has_data:
            rows_written += 1
        dest_row += 1
    data_last_row = 7 + rows_written

    for tbl_name in ws_data_tab.tables:
        tbl = ws_data_tab.tables[tbl_name]
        if tbl.displayName == "Table1":
            tbl.ref = f"B7:X{data_last_row}"
            break

    for r in range(8, data_last_row + 1):
        ws_data_tab[f"B{r}"] = "=Table1[[#This Row],[VendorID2]]"
        ws_data_tab[f"C{r}"] = "=Table1[[#This Row],[Vendor Name]]"
        ws_data_tab[f"D{r}"] = "=_xlfn.XLOOKUP(Table1[[#This Row],[Original Vendor]],'Vendor Dedupe'!B:B,'Vendor Dedupe'!D:D,\"\")"
        ws_data_tab[f"E{r}"] = "=Table1[[#This Row],[Invoice Description3]]"
        ws_data_tab[f"F{r}"] = "=Table1[[#This Row],[GL Description2]]"
        ws_data_tab[f"G{r}"] = "=Table1[[#This Row],[Total amount]]"
        ws_data_tab[f"H{r}"] = "=Table1[[#This Row],[Date2]]"
        ws_data_tab[f"I{r}"] = "=_xlfn.XLOOKUP(Table1[[#This Row],[Deduped Vendor]],'Vendor Map'!B:B,'Vendor Map'!E:E,\"\")"
        ws_data_tab[f"J{r}"] = "=Table1[[#This Row],[GL Description2]]"
        ws_data_tab[f"K{r}"] = "=_xlfn.XLOOKUP(Table1[[#This Row],[Original Vendor]],'Vendor Map'!B:B,'Vendor Map'!D:D,\"\")"
        ws_data_tab[f"L{r}"] = (
            f"=IFERROR(IF(OR(I{r}=\"\",I{r}=0),IFERROR(IF(OR(J{r}=\"\",J{r}=0),\"Unmapped\",J{r}),\"Unmapped\"),I{r}),"
            f"IFERROR(IF(OR(J{r}=\"\",J{r}=0),\"Unmapped\",J{r}),\"Unmapped\"))"
        )
        ws_data_tab[f"M{r}"] = "=Table1[[#This Row],[Source]]"
    set_sheet_view(ws_data_tab)
    status(f"  ✅ Data tab: {rows_written} rows.")

    # ----- STEP 7: Spend summaries -----
    category_spend = defaultdict(float)
    area_by_category = {}
    vendor_spend_by_std = defaultdict(float)
    category_by_vendor = {}
    for raw_vendor in vendor_rows:
        spend = vendor_spend.get(raw_vendor, 0)
        canonical = get_canonical(raw_vendor)
        std_name = vendor_memory.get(canonical, {}).get("standard_name", canonical)
        mapping = vendor_map_memory.get(std_name, {})
        category = mapping.get("category_mapping", "Unmapped")
        area = mapping.get("area_mapping", "Unaddressable")
        category_spend[category] += spend
        area_by_category[category] = area
        vendor_spend_by_std[std_name] += spend
        category_by_vendor[std_name] = category

    # ----- STEP 8: Spend by Category tab -----
    status("\n📊 Populating Spend by Category…")
    ws_sbc._pivots = []
    for row in ws_sbc.iter_rows(min_row=4, max_row=ws_sbc.max_row):
        for cell in row:
            cell.value = None

    sorted_cats = sorted(category_spend.items(), key=lambda x: x[1], reverse=True)
    sorted_vendors = sorted(vendor_spend_by_std.items(), key=lambda x: x[1], reverse=True)
    cat_last_row = 5 + len(sorted_cats)
    vendor_last_row = 5 + len(sorted_vendors)

    ws_sbc["E4"] = f"=SUBTOTAL(9,E6:E{cat_last_row})"
    ws_sbc["E4"].font = Font(name="Calibri", bold=True, size=10)
    ws_sbc["I4"] = f"=SUBTOTAL(9,I6:I{vendor_last_row})"
    ws_sbc["I4"].font = Font(name="Calibri", bold=True, size=10)
    for col, hdr in {2: "Final Mapping", 3: "Final Spend Area", 4: "Deduped Vendor", 5: "Spend",
                     7: "Deduped Vendor", 8: "Final Mapping", 9: "Spend"}.items():
        cell = ws_sbc.cell(row=5, column=col)
        cell.value = hdr
        style_table_header(cell)

    for i, (cat, _) in enumerate(sorted_cats):
        r = 6 + i
        ws_sbc.cell(row=r, column=2).value = cat
        ws_sbc.cell(row=r, column=3).value = area_by_category.get(cat, "")
        ws_sbc.cell(row=r, column=5).value = f"=SUMIFS(Data!$G:$G,Data!$L:$L,B{r})"
        style_data_row(ws_sbc.cell(row=r, column=2))
        style_data_row(ws_sbc.cell(row=r, column=3))
        style_data_row(ws_sbc.cell(row=r, column=5))
    gt1_row = 6 + len(sorted_cats)
    ws_sbc.cell(row=gt1_row, column=2).value = "Grand Total"
    ws_sbc.cell(row=gt1_row, column=5).value = f"=SUM(E6:E{gt1_row - 1})"
    style_total_row(ws_sbc.cell(row=gt1_row, column=2))
    style_total_row(ws_sbc.cell(row=gt1_row, column=5))

    for i, (vendor, _) in enumerate(sorted_vendors):
        r = 6 + i
        ws_sbc.cell(row=r, column=7).value = vendor
        ws_sbc.cell(row=r, column=8).value = category_by_vendor.get(vendor, "")
        ws_sbc.cell(row=r, column=9).value = f"=SUMIFS(Data!$G:$G,Data!$D:$D,G{r})"
        style_data_row(ws_sbc.cell(row=r, column=7))
        style_data_row(ws_sbc.cell(row=r, column=8))
        style_data_row(ws_sbc.cell(row=r, column=9))
    gt2_row = 6 + len(sorted_vendors)
    ws_sbc.cell(row=gt2_row, column=7).value = "Grand Total"
    ws_sbc.cell(row=gt2_row, column=9).value = f"=SUM(I6:I{gt2_row - 1})"
    style_total_row(ws_sbc.cell(row=gt2_row, column=7))
    style_total_row(ws_sbc.cell(row=gt2_row, column=9))

    ws_sbc.column_dimensions["A"].width = 2.63
    ws_sbc.column_dimensions["B"].width = 25.63
    ws_sbc.column_dimensions["C"].width = 25.63
    ws_sbc.column_dimensions["D"].width = 25.63
    ws_sbc.column_dimensions["F"].width = 2.63
    ws_sbc.column_dimensions["G"].width = 25.63
    set_sheet_view(ws_sbc)

    # ----- STEP 9: Select top categories -----
    top_categories = _select_top_categories(sorted_cats, status)

    # ----- STEP 10: Savings benchmarks -----
    status("\n💡 Getting savings benchmarks…")
    benchmarks, new_benchmarks = get_savings_benchmarks(
        client, [cat for cat, _ in top_categories], research_summary,
        savings_cache, company_name, industry, pe_firm, status,
    )
    db.save_savings_benchmarks(supabase, new_benchmarks)

    # ----- STEP 11: Savings Opportunity tab -----
    status("\n📈 Populating Savings Opportunity…")
    ws_sav["L4"] = "Internal Notes"
    ws_sav["L4"].font = Font(name="Calibri", bold=True, size=11, color="000000")
    n_categories = len(top_categories)
    template_data_rows = 13
    first_data_row = 5
    blank_gap_row = 18
    template_total_row = 19
    if n_categories > template_data_rows:
        rows_to_insert = n_categories - template_data_rows
        ws_sav.insert_rows(blank_gap_row, amount=rows_to_insert)
        total_row = template_total_row + rows_to_insert
    else:
        total_row = template_total_row
    last_data_row = first_data_row + n_categories - 1

    for r in range(first_data_row, first_data_row + template_data_rows + 1):
        for col in [3, 5, 6, 7, 12]:
            ws_sav.cell(row=r, column=col).value = None

    for i, (cat, _) in enumerate(top_categories):
        r = first_data_row + i
        bm = benchmarks.get(cat, {"addressability": 0.8, "savings_low_pct": 0.01,
                                  "savings_high_pct": 0.02, "notes": "Default benchmark — please review"})
        ws_sav.cell(row=r, column=3).value = cat
        ws_sav.cell(row=r, column=4).value = f"=SUMIFS(Data!$G:$G,Data!$L:$L,C{r})"
        ws_sav.cell(row=r, column=5).value = bm["addressability"]
        ws_sav.cell(row=r, column=6).value = bm["savings_low_pct"]
        ws_sav.cell(row=r, column=7).value = bm["savings_high_pct"]
        ws_sav.cell(row=r, column=8).value = f"=F{r}*D{r}*E{r}"
        ws_sav.cell(row=r, column=9).value = f"=G{r}*E{r}*D{r}"
        ws_sav.cell(row=r, column=10).value = f"=AVERAGE(H{r}:I{r})"
        ws_sav.cell(row=r, column=12).value = bm["notes"]
    ws_sav.cell(row=total_row, column=3).value = "TOTAL"
    ws_sav.cell(row=total_row, column=4).value = f"=SUM(D{first_data_row}:D{last_data_row})"
    ws_sav.cell(row=total_row, column=8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws_sav.cell(row=total_row, column=9).value = f"=SUM(I{first_data_row}:I{last_data_row})"
    ws_sav.cell(row=total_row, column=10).value = f"=SUM(J{first_data_row}:J{last_data_row})"
    set_sheet_view(ws_sav)

    # ----- STEP 12: QC sheet -----
    if "QC" in wb.sheetnames:
        del wb["QC"]
    ws_qc = wb.create_sheet("QC")
    set_sheet_view(ws_qc)
    ws_qc["B2"] = company_name
    style_heading(ws_qc["B2"], size=14, bold=True)
    ws_qc["B3"] = "Spend and Opportunity Assessment"
    style_heading(ws_qc["B3"], size=11, bold=False, color=MID_GREY)
    ws_qc["B4"] = "QC — Control Total Check"
    style_heading(ws_qc["B4"], size=11, bold=True)
    ws_qc["C6"] = f"Run Date: {datetime.today().strftime('%Y-%m-%d %H:%M')}"
    style_body(ws_qc["C6"])
    ws_qc["C7"] = "All four spend totals below should match when the file is opened in Excel."
    ws_qc["C7"].font = Font(name="Calibri", bold=False, size=10, color="000000")
    ws_qc["C7"].alignment = Alignment(wrap_text=True, vertical="top")
    for col, hdr in {3: "Source", 4: "Tab", 5: "Cell Reference", 6: "Spend Amount", 7: "Status"}.items():
        cell = ws_qc.cell(row=9, column=col)
        cell.value = hdr
        style_table_header(cell)
    for i, (source, tab, ref) in enumerate([
        ("Raw Data Total", "Consolidated Data", "I4"),
        ("Data Tab Total", "Data", "G5"),
        ("Vendor Map Total", "Vendor Map", "C4"),
        ("Vendor Dedupe Total", "Vendor Dedupe", "C4"),
    ], start=10):
        ws_qc.cell(row=i, column=3).value = source
        ws_qc.cell(row=i, column=4).value = tab
        ws_qc.cell(row=i, column=5).value = ref
        ws_qc.cell(row=i, column=6).value = round(total_raw_spend, 2)
        ws_qc.cell(row=i, column=7).value = "Verify in Excel after opening"
        ws_qc.cell(row=i, column=7).font = Font(name="Calibri", bold=False, size=10, color="007700")
        for col in [3, 4, 5, 6]:
            style_data_row(ws_qc.cell(row=i, column=col))
    ws_qc.cell(row=15, column=3).value = "OVERALL STATUS:"
    style_subheading(ws_qc.cell(row=15, column=3))
    ws_qc.cell(row=15, column=4).value = f"Total Spend = ${total_raw_spend:,.2f} — All four tabs should match."
    ws_qc.cell(row=15, column=4).font = Font(name="Calibri", bold=True, size=10, color="007700")
    ws_qc.column_dimensions["A"].width = 2.63
    ws_qc.column_dimensions["B"].width = 3
    ws_qc.column_dimensions["C"].width = 25
    ws_qc.column_dimensions["D"].width = 20
    ws_qc.column_dimensions["E"].width = 18
    ws_qc.column_dimensions["F"].width = 20
    ws_qc.column_dimensions["G"].width = 40

    # ----- Build dashboard data (IRONCLAD shape for the React dashboard) -----
    import dashboard_export
    dashboard_data = dashboard_export.build_ironclad_data(
        ws_cd=ws_cd,
        cd_last_row=cd_last_row,
        company_name=company_name,
        pe_firm=pe_firm,
        industry=industry,
        research_summary=research_summary,
        unique_vendors=unique_vendors,
        get_canonical=get_canonical,
        vendor_memory=vendor_memory,
        vendor_map_memory=vendor_map_memory,
        vendor_spend=vendor_spend,
        benchmarks=benchmarks,
        total_raw_spend=total_raw_spend,
        status=status,
    )

    # ----- Save to bytes -----
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    today_date = datetime.today().strftime("%Y-%m-%d")
    suggested_filename = f"{company_name}_AP_Analysis_{today_date}.xlsx"

    status(f"\n✅ Done. Total spend analysed: ${total_raw_spend:,.2f}")
    return out_buf.getvalue(), suggested_filename, dashboard_data
